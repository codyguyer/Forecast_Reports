from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(r"c:\Users\cguyer\OneDrive - Midmark Corporation\Documents\Sales Ops\Reporting\Ad Hoc Reports\Forecast Reports")

MARKETING_FILE = ROOT / "Marketing Forecast Data.xlsx"
PRODUCT_CATALOG_FILE = ROOT / "product_catalog_master.xlsx"
ACTUALS_FILE = ROOT / "all_products_actuals_and_bookings.xlsx"


@dataclass
class ReportConfig:
    report_month: date
    marketing_file: Path
    product_catalog_file: Path
    stats_model_file: Path
    output_file: Path


def first_day_of_month(d: date) -> date:
    return date(d.year, d.month, 1)


def previous_month(d: date) -> date:
    if d.month == 1:
        return date(d.year - 1, 12, 1)
    return date(d.year, d.month - 1, 1)


def month_label(d: date) -> str:
    return d.strftime("%b")


def stats_model_filename(d: date) -> str:
    return f"stats_model_forecasts_{d.year}-{d.strftime('%b')}.xlsx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build forecast accuracy report.")
    parser.add_argument(
        "--month",
        type=str,
        default=None,
        help="Report month in YYYY-MM (defaults to previous month of today).",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output XLSX filename (optional).",
    )
    return parser.parse_args()


def resolve_report_month(arg: Optional[str]) -> date:
    if not arg:
        return previous_month(date.today())
    try:
        parsed = datetime.strptime(arg, "%Y-%m").date()
    except ValueError:
        raise SystemExit("--month must be in YYYY-MM format")
    return first_day_of_month(parsed)


def load_marketing_data(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Tableau Data Pull")
    df = df.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    return df

def load_actuals_data(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    df = df.copy()
    df["Month"] = pd.to_datetime(df["Month"], errors="coerce")
    return df


def load_product_catalog(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    df = df.copy()
    # Override any BU labeled "Division" to D200 for consistency
    df["business_unit_code"] = df["business_unit_code"].replace({"Division": "D200"})
    df["business_unit_name"] = df["business_unit_name"].replace({"Division": "D200"})
    return df


def load_stats_model(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Forecast_Library")
    df = df.copy()
    df["forecast_month"] = pd.to_datetime(df["forecast_month"], errors="coerce")
    return df


def normalize_key(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.upper()


def build_raw_data(
    marketing_df: pd.DataFrame,
    catalog_df: pd.DataFrame,
    stats_df: pd.DataFrame,
    actuals_df: pd.DataFrame,
    report_month: date,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    month_start = pd.Timestamp(report_month)

    marketing_df = marketing_df.copy()
    marketing_df = marketing_df[marketing_df["Date"].dt.to_period("M") == month_start.to_period("M")]
    marketing_df = marketing_df[marketing_df["Geography"].astype(str).str.upper() == "AMERICAS"]

    catalog_df = catalog_df.copy()

    # Expand sku_list to allow multiple products per catalog row
    catalog_map = catalog_df.copy()
    catalog_map["sku_list"] = catalog_map["sku_list"].astype(str).str.split("|")
    catalog_map = catalog_map.explode("sku_list")
    catalog_map["sku_list"] = catalog_map["sku_list"].astype(str).str.strip()

    marketing_df["_product_key"] = normalize_key(marketing_df["Product"])
    marketing_df["_bu_key"] = normalize_key(marketing_df["BU"])

    catalog_map["_product_key"] = normalize_key(catalog_map["sku_list"])
    catalog_map["_bu_key"] = normalize_key(catalog_map["business_unit_code"])

    merged = marketing_df.merge(
        catalog_map,
        on=["_product_key", "_bu_key"],
        how="inner",
        suffixes=("", "_catalog"),
    )

    # Casework split for D200 based on Location
    casework_location_map = {
        "LOC1020": "ARTISAN CASEWORK",
        "LOC1080": "SYNTHESIS CASEWORK",
    }
    merged["_location_key"] = normalize_key(merged["Location"])
    merged["_group_key_norm"] = normalize_key(merged["group_key"])
    casework_mask = (
        (merged["_bu_key"] == "D200")
        & (merged["_product_key"] == "TOTAL CASEWORK")
        & (merged["_location_key"].isin(casework_location_map.keys()))
    )
    if casework_mask.any():
        mapped_group = merged["_location_key"].map(casework_location_map)
        merged = merged[~casework_mask | (merged["_group_key_norm"] == mapped_group)]

    skipped = marketing_df.merge(
        catalog_map[["_product_key", "_bu_key"]],
        on=["_product_key", "_bu_key"],
        how="left",
        indicator=True,
    )
    skipped = skipped[skipped["_merge"] == "left_only"].copy()
    skipped["Reason"] = "Not in product catalog master"

    merged["Units"] = merged["salesforce_feature_mode"].astype(str).str.lower().str.strip()

    merged["Marketing Fcast"] = merged.apply(
        lambda r: r["Forecast (Quantity)"] if r["Units"] == "quantity" else r["Forecast (Dollars)"],
        axis=1,
    )

    merged["Marketing Fcast"] = pd.to_numeric(merged["Marketing Fcast"], errors="coerce").fillna(0)

    agg = (
        merged.groupby(
            [
                "group_key",
                "business_unit_code",
                "business_unit_name",
                "product_family",
                "marketing_manager",
                "Units",
            ],
            dropna=False,
        )
        .agg(
            Marketing_Fcast=("Marketing Fcast", "sum"),
        )
        .reset_index()
    )

    # Actuals rollup from separate file (by group_key + BU + month)
    actuals_df = actuals_df.copy()
    actuals_df = actuals_df[actuals_df["Month"].dt.to_period("M") == month_start.to_period("M")]
    actuals_df["_product_key"] = normalize_key(actuals_df["Product"])
    actuals_df["_bu_key"] = normalize_key(actuals_df["Division"])
    actuals_rollup = (
        actuals_df.groupby(["_product_key", "_bu_key"], dropna=False)["Actuals"]
        .sum()
        .reset_index()
    )

    stats_df = stats_df.copy()
    stats_df = stats_df[stats_df["forecast_month"].dt.to_period("M") == month_start.to_period("M")]
    stats_df["_product_key"] = normalize_key(stats_df["product_id"])
    stats_df["_bu_key"] = normalize_key(stats_df["bu_id"])
    stats_df["model_type"] = stats_df["model_type"].astype(str).str.upper().str.strip()

    # Prefer BLEND when available, otherwise fall back to recommended_model == True
    stats_blend = stats_df[stats_df["model_type"] == "BLEND"].copy()
    stats_fallback = stats_df[stats_df["recommended_model"] == True].copy()

    stats_rollup_blend = (
        stats_blend.groupby(["_product_key", "_bu_key"], dropna=False)["forecast_value"]
        .sum()
        .reset_index()
    )
    stats_rollup_fallback = (
        stats_fallback.groupby(["_product_key", "_bu_key"], dropna=False)["forecast_value"]
        .sum()
        .reset_index()
    )

    stats_rollup = stats_rollup_blend.merge(
        stats_rollup_fallback,
        on=["_product_key", "_bu_key"],
        how="outer",
        suffixes=("_blend", "_fallback"),
    )
    stats_rollup["forecast_value"] = stats_rollup["forecast_value_blend"].where(
        stats_rollup["forecast_value_blend"].notna(),
        stats_rollup["forecast_value_fallback"],
    )
    stats_rollup = stats_rollup[["_product_key", "_bu_key", "forecast_value"]]

    agg["_product_key"] = normalize_key(agg["group_key"])
    agg["_bu_key"] = normalize_key(agg["business_unit_code"])

    agg = agg.merge(
        actuals_rollup,
        on=["_product_key", "_bu_key"],
        how="left",
    )

    agg = agg.merge(
        stats_rollup,
        on=["_product_key", "_bu_key"],
        how="left",
    )

    agg["Stats Model Fcast"] = pd.to_numeric(agg["forecast_value"], errors="coerce").fillna(0)
    agg["Marketing_Fcast"] = pd.to_numeric(agg["Marketing_Fcast"], errors="coerce").fillna(0)
    agg["Actuals"] = pd.to_numeric(agg["Actuals"], errors="coerce").fillna(0)

    agg["Stats Abs Error"] = (agg["Actuals"] - agg["Stats Model Fcast"]).abs()
    agg["Marketing Abs Error"] = (agg["Actuals"] - agg["Marketing_Fcast"]).abs()

    raw = pd.DataFrame(
        {
            "Product": agg["group_key"],
            "Division": agg["business_unit_code"],
            "Month": month_start,
            "sku": "",  # rollup level
            "Prod Fam": agg["product_family"],
            "Units": agg["Units"],
            "Actuals": agg["Actuals"],
            "Stats Model Fcast": agg["Stats Model Fcast"],
            "Marketing Fcast": agg["Marketing_Fcast"],
            "Stats Abs Error": agg["Stats Abs Error"],
            "Marketing Abs Error": agg["Marketing Abs Error"],
            "BU Name": agg["business_unit_name"],
            "Marketing Manager": agg["marketing_manager"],
        }
    )

    skipped_out = skipped[["BU", "Product", "Date", "Reason"]].copy()
    skipped_out = skipped_out.sort_values(["BU", "Product"]) if not skipped_out.empty else skipped_out

    marketing_detail = merged[
        [
            "BU",
            "Product",
            "Date",
            "group_key",
            "business_unit_code",
            "business_unit_name",
            "product_family",
            "marketing_manager",
            "Units",
            "Marketing Fcast",
        ]
    ].copy()

    return raw, skipped_out, marketing_detail, stats_rollup, agg, actuals_rollup


def safe_ratio(numerator: float, denominator: float) -> Optional[float]:
    if denominator == 0:
        return None
    return numerator / denominator


def build_totals_dashboard(raw: pd.DataFrame) -> pd.DataFrame:
    records = []

    for units in ["quantity", "dollars"]:
        unit_df = raw[raw["Units"] == units]
        if unit_df.empty:
            continue

        for (bu_code, bu_name), group in (
            unit_df.groupby(["Division", "BU Name"], dropna=False)
        ):
            actuals_sum = group["Actuals"].sum()
            stats_sum = group["Stats Model Fcast"].sum()
            mkt_sum = group["Marketing Fcast"].sum()
            stats_abs = group["Stats Abs Error"].sum()
            mkt_abs = group["Marketing Abs Error"].sum()

            stats_acc = safe_ratio(stats_sum, actuals_sum)
            mkt_acc = safe_ratio(mkt_sum, actuals_sum)
            stats_wape = safe_ratio(stats_abs, actuals_sum)
            mkt_wape = safe_ratio(mkt_abs, actuals_sum)

            records.append(
                {
                    "Units": units,
                    "Scope": "BU",
                    "BU Code": bu_code,
                    "BU Name": bu_name,
                    "Metric": "Midmark Fcast Acc.",
                    "Stats Model": stats_acc,
                    "Marketing": mkt_acc,
                }
            )
            records.append(
                {
                    "Units": units,
                    "Scope": "BU",
                    "BU Code": bu_code,
                    "BU Name": bu_name,
                    "Metric": "WAPE",
                    "Stats Model": stats_wape,
                    "Marketing": mkt_wape,
                }
            )

        # total across all BUs for unit
        actuals_sum = unit_df["Actuals"].sum()
        stats_sum = unit_df["Stats Model Fcast"].sum()
        mkt_sum = unit_df["Marketing Fcast"].sum()
        stats_abs = unit_df["Stats Abs Error"].sum()
        mkt_abs = unit_df["Marketing Abs Error"].sum()

        stats_acc = safe_ratio(stats_sum, actuals_sum)
        mkt_acc = safe_ratio(mkt_sum, actuals_sum)
        stats_wape = safe_ratio(stats_abs, actuals_sum)
        mkt_wape = safe_ratio(mkt_abs, actuals_sum)

        records.append(
            {
                "Units": units,
                "Scope": "Total",
                "BU Code": "ALL",
                "BU Name": "Total",
                "Metric": "Midmark Fcast Acc.",
                "Stats Model": stats_acc,
                "Marketing": mkt_acc,
            }
        )
        records.append(
            {
                "Units": units,
                "Scope": "Total",
                "BU Code": "ALL",
                "BU Name": "Total",
                "Metric": "WAPE",
                "Stats Model": stats_wape,
                "Marketing": mkt_wape,
            }
        )

    return pd.DataFrame.from_records(records)


def build_prod_fam_dashboard(raw: pd.DataFrame) -> pd.DataFrame:
    records = []
    for units in ["quantity", "dollars"]:
        unit_df = raw[raw["Units"] == units]
        if unit_df.empty:
            continue

        grouped = unit_df.groupby(["Division", "BU Name", "Prod Fam"], dropna=False)
        for (bu_code, bu_name, prod_fam), group in grouped:
            actuals_sum = group["Actuals"].sum()
            stats_sum = group["Stats Model Fcast"].sum()
            mkt_sum = group["Marketing Fcast"].sum()
            stats_abs = group["Stats Abs Error"].sum()
            mkt_abs = group["Marketing Abs Error"].sum()

            stats_acc = safe_ratio(stats_sum, actuals_sum)
            mkt_acc = safe_ratio(mkt_sum, actuals_sum)
            stats_wape = safe_ratio(stats_abs, actuals_sum)
            mkt_wape = safe_ratio(mkt_abs, actuals_sum)

            records.append(
                {
                    "Units": units,
                    "BU Code": bu_code,
                    "BU Name": bu_name,
                    "Prod Fam": prod_fam,
                    "Metric": "Midmark Fcast Acc.",
                    "Stats Model": stats_acc,
                    "Marketing": mkt_acc,
                }
            )

    return pd.DataFrame.from_records(records)

def build_prod_fam_wape_dashboard(raw: pd.DataFrame) -> pd.DataFrame:
    records = []
    for units in ["quantity", "dollars"]:
        unit_df = raw[raw["Units"] == units]
        if unit_df.empty:
            continue

        grouped = unit_df.groupby(["Division", "BU Name", "Prod Fam"], dropna=False)
        for (bu_code, bu_name, prod_fam), group in grouped:
            actuals_sum = group["Actuals"].sum()
            stats_sum = group["Stats Model Fcast"].sum()
            mkt_sum = group["Marketing Fcast"].sum()
            stats_abs = group["Stats Abs Error"].sum()
            mkt_abs = group["Marketing Abs Error"].sum()

            stats_wape = safe_ratio(stats_abs, actuals_sum)
            mkt_wape = safe_ratio(mkt_abs, actuals_sum)

            records.append(
                {
                    "Units": units,
                    "BU Code": bu_code,
                    "BU Name": bu_name,
                    "Prod Fam": prod_fam,
                    "Metric": "WAPE",
                    "Stats Model": stats_wape,
                    "Marketing": mkt_wape,
                }
            )

    return pd.DataFrame.from_records(records)

def build_marketing_manager_dashboard(raw: pd.DataFrame) -> pd.DataFrame:
    records = []
    for units in ["quantity", "dollars"]:
        unit_df = raw[raw["Units"] == units]
        if unit_df.empty:
            continue

        grouped = unit_df.groupby(["Division", "BU Name", "Marketing Manager"], dropna=False)
        for (bu_code, bu_name, manager), group in grouped:
            actuals_sum = group["Actuals"].sum()
            stats_sum = group["Stats Model Fcast"].sum()
            mkt_sum = group["Marketing Fcast"].sum()
            stats_abs = group["Stats Abs Error"].sum()
            mkt_abs = group["Marketing Abs Error"].sum()

            stats_acc = safe_ratio(stats_sum, actuals_sum)
            mkt_acc = safe_ratio(mkt_sum, actuals_sum)
            stats_wape = safe_ratio(stats_abs, actuals_sum)
            mkt_wape = safe_ratio(mkt_abs, actuals_sum)
            stats_wape_acc = None if stats_wape is None else 1 - stats_wape
            mkt_wape_acc = None if mkt_wape is None else 1 - mkt_wape

            records.append(
                {
                    "Units": units,
                    "BU Code": bu_code,
                    "BU Name": bu_name,
                    "Marketing Manager": manager,
                    "Metric": "Midmark Fcast Acc.",
                    "Stats Model": stats_acc,
                    "Marketing": mkt_acc,
                }
            )
            records.append(
                {
                    "Units": units,
                    "BU Code": bu_code,
                    "BU Name": bu_name,
                    "Marketing Manager": manager,
                    "Metric": "WAPE",
                    "Stats Model": stats_wape,
                    "Marketing": mkt_wape,
                }
            )

    return pd.DataFrame.from_records(records)

def build_product_dashboard(raw: pd.DataFrame) -> pd.DataFrame:
    records = []
    for units in ["quantity", "dollars"]:
        unit_df = raw[raw["Units"] == units]
        if unit_df.empty:
            continue

        grouped = unit_df.groupby(
            ["Division", "BU Name", "Product", "Marketing Manager"], dropna=False
        )
        for (bu_code, bu_name, product, manager), group in grouped:
            actuals_sum = group["Actuals"].sum()
            stats_sum = group["Stats Model Fcast"].sum()
            mkt_sum = group["Marketing Fcast"].sum()

            stats_acc = safe_ratio(stats_sum, actuals_sum)
            mkt_acc = safe_ratio(mkt_sum, actuals_sum)

            records.append(
                {
                    "Units": units,
                    "BU Code": bu_code,
                    "BU Name": bu_name,
                    "Product": product,
                    "Marketing Manager": manager,
                    "Metric": "Midmark Fcast Acc.",
                    "Stats Model": stats_acc,
                    "Marketing": mkt_acc,
                }
            )

    return pd.DataFrame.from_records(records)


def write_report(config: ReportConfig) -> None:
    marketing_df = load_marketing_data(config.marketing_file)
    catalog_df = load_product_catalog(config.product_catalog_file)
    stats_df = load_stats_model(config.stats_model_file)
    actuals_df = load_actuals_data(ACTUALS_FILE)

    raw, skipped, marketing_detail, stats_rollup, marketing_rollup, actuals_rollup = build_raw_data(
        marketing_df, catalog_df, stats_df, actuals_df, config.report_month
    )
    totals = build_totals_dashboard(raw)
    prod_fam = build_prod_fam_dashboard(raw)
    prod_fam_wape = build_prod_fam_wape_dashboard(raw)
    marketing_mgr = build_marketing_manager_dashboard(raw)
    product = build_product_dashboard(raw)

    with pd.ExcelWriter(config.output_file, engine="openpyxl") as writer:
        totals.to_excel(writer, sheet_name="Dashboard - Totals Data", index=False)
        prod_fam.to_excel(writer, sheet_name="Dashboard - Prod Fam Data", index=False)
        prod_fam_wape.to_excel(writer, sheet_name="Dashboard - Prod Fam WAPE Data", index=False)
        marketing_mgr.to_excel(writer, sheet_name="Dashboard - Mkt Mgr Data", index=False)
        product.to_excel(writer, sheet_name="Dashboard - Product Data", index=False)
        raw.to_excel(writer, sheet_name="Raw Data", index=False)
        skipped.to_excel(writer, sheet_name="Skipped Products", index=False)
        catalog_df.to_excel(writer, sheet_name="Lookup", index=False)

    wb = load_workbook(config.output_file)
    highlight = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=13)
    percent_fmt = "0.0%"
    thin = Side(style="thin", color="A0A0A0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_section_header(ws, row, title, col_span=5):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = title_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="left")

    def write_table_header(ws, row, headers):
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    def write_metric_row(ws, row, values, percent_cols):
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = border
            if col_idx in percent_cols:
                cell.number_format = percent_fmt
            cell.alignment = Alignment(horizontal="center")

    def highlight_winner(ws, row, stats_col, mkt_col, mode="closest_to_one"):
        stats_val = ws.cell(row=row, column=stats_col).value
        mkt_val = ws.cell(row=row, column=mkt_col).value
        if stats_val is None or mkt_val is None:
            return
        try:
            stats_val = float(stats_val)
            mkt_val = float(mkt_val)
        except (TypeError, ValueError):
            return
        if stats_val == mkt_val:
            return
        if mode == "higher":
            if stats_val > mkt_val:
                ws.cell(row=row, column=stats_col).fill = highlight
            else:
                ws.cell(row=row, column=mkt_col).fill = highlight
        elif mode == "lower":
            if stats_val < mkt_val:
                ws.cell(row=row, column=stats_col).fill = highlight
            else:
                ws.cell(row=row, column=mkt_col).fill = highlight
        else:
            stats_diff = abs(stats_val - 1)
            mkt_diff = abs(mkt_val - 1)
            if stats_diff == mkt_diff:
                return
            if stats_diff < mkt_diff:
                ws.cell(row=row, column=stats_col).fill = highlight
            else:
                ws.cell(row=row, column=mkt_col).fill = highlight

    def winner_label(stats_val, mkt_val, mode="closest_to_one"):
        if stats_val is None or mkt_val is None:
            return ""
        if pd.isna(stats_val) or pd.isna(mkt_val):
            return ""
        try:
            stats_val = float(stats_val)
            mkt_val = float(mkt_val)
        except (TypeError, ValueError):
            return ""
        if stats_val == mkt_val:
            return "Tie"
        if mode == "higher":
            return "Stats Model" if stats_val > mkt_val else "Marketing"
        if mode == "lower":
            return "Stats Model" if stats_val < mkt_val else "Marketing"
        stats_diff = abs(stats_val - 1)
        mkt_diff = abs(mkt_val - 1)
        if stats_diff == mkt_diff:
            return "Tie"
        return "Stats Model" if stats_diff < mkt_diff else "Marketing"

    def build_dashboard_totals(ws_name: str, df: pd.DataFrame) -> None:
        if ws_name in wb.sheetnames:
            del wb[ws_name]
        ws = wb.create_sheet(ws_name, 0)
        ws.title = ws_name
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

        row = 1
        ws.cell(row=row, column=1, value="Forecast Accuracy Dashboard - Totals").font = title_font
        row += 2

        for scope in ["Total", "BU"]:
            scope_df = df[df["Scope"] == scope]
            if scope_df.empty:
                continue
            if scope == "Total":
                write_section_header(ws, row, "Totals", col_span=3)
                row += 2
                bu_groups = [("ALL", "Total")]
            else:
                write_section_header(ws, row, "By Business Unit", col_span=3)
                row += 2
                bu_order = {"D100": 0, "D200": 1, "D300": 2}
                bu_groups = (
                    scope_df[["BU Code", "BU Name"]]
                    .drop_duplicates()
                    .assign(_order=lambda d: d["BU Code"].map(bu_order).fillna(99))
                    .sort_values(["_order", "BU Code"])
                    .drop(columns=["_order"])
                    .itertuples(index=False, name=None)
                )

            for bu_code, bu_name in bu_groups:
                bu_df = scope_df[(scope_df["BU Code"] == bu_code) & (scope_df["BU Name"] == bu_name)]
                if bu_df.empty:
                    continue
                ws.cell(row=row, column=1, value=f"{bu_name} ({bu_code})").font = header_font
                row += 1
                for units in ["quantity", "dollars"]:
                    unit_df = bu_df[bu_df["Units"] == units]
                    if unit_df.empty:
                        continue
                    ws.cell(row=row, column=1, value=units.title()).font = header_font
                    row += 1
                    write_table_header(ws, row, ["Metric", "Stats Model", "Marketing"])
                    row += 1
                    for metric in ["Midmark Fcast Acc.", "WAPE"]:
                        metric_df = unit_df[unit_df["Metric"] == metric]
                        if metric_df.empty:
                            continue
                        stats_val = metric_df["Stats Model"].iloc[0]
                        mkt_val = metric_df["Marketing"].iloc[0]
                        write_metric_row(ws, row, [metric, stats_val, mkt_val], percent_cols={2, 3})
                        highlight_winner(ws, row, 2, 3, mode=("lower" if metric == "WAPE" else "closest_to_one"))
                        if metric == "WAPE":
                            ws.row_dimensions[row].hidden = True
                        row += 1
                    row += 1
                row += 1

        for r in range(1, row):
            ws.row_dimensions[r].height = 18

    def build_dashboard_prod_fam(ws_name: str, df: pd.DataFrame) -> None:
        if ws_name in wb.sheetnames:
            del wb[ws_name]
        ws = wb.create_sheet(ws_name, 1)
        ws.title = ws_name
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["B"].hidden = True

        row = 1
        ws.cell(row=row, column=1, value="Forecast Accuracy Dashboard - Product Family").font = title_font
        row += 2

        bu_order = {"D100": 0, "D200": 1, "D300": 2}
        for (bu_code, bu_name) in (
            df[["BU Code", "BU Name"]]
            .drop_duplicates()
            .assign(_order=lambda d: d["BU Code"].map(bu_order).fillna(99))
            .sort_values(["_order", "BU Code"])
            .drop(columns=["_order"])
            .itertuples(index=False, name=None)
        ):
            bu_df = df[(df["BU Code"] == bu_code) & (df["BU Name"] == bu_name)]
            if bu_df.empty:
                continue
            write_section_header(ws, row, f"{bu_name} ({bu_code})", col_span=4)
            row += 2
            for units in ["quantity", "dollars"]:
                unit_df = bu_df[bu_df["Units"] == units]
                if unit_df.empty:
                    continue
                ws.cell(row=row, column=1, value=units.title()).font = header_font
                row += 1
                write_table_header(ws, row, ["Prod Fam", "Metric", "Stats Model", "Marketing"])
                row += 1
                prod_fams = unit_df["Prod Fam"].dropna().unique().tolist()
                sort_rows = []
                for prod_fam in prod_fams:
                    pf_df = unit_df[(unit_df["Prod Fam"] == prod_fam) & (unit_df["Metric"] == "Midmark Fcast Acc.")]
                    if pf_df.empty:
                        continue
                    stats_val = pf_df["Stats Model"].iloc[0]
                    mkt_val = pf_df["Marketing"].iloc[0]
                    winner = winner_label(stats_val, mkt_val, mode="closest_to_one")
                    winner_rank = 2
                    if winner == "Stats Model":
                        winner_rank = 0
                    elif winner == "Marketing":
                        winner_rank = 1
                    prod_fam_sort = str(prod_fam).strip().lower()
                    sort_rows.append((winner_rank, prod_fam_sort, str(prod_fam), stats_val, mkt_val))

                for _, _, prod_fam, stats_val, mkt_val in sorted(
                    sort_rows, key=lambda r: (r[0], r[1])
                ):
                    write_metric_row(
                        ws,
                        row,
                        [prod_fam, "Midmark Fcast Acc.", stats_val, mkt_val],
                        percent_cols={3, 4},
                    )
                    highlight_winner(ws, row, 3, 4, mode="closest_to_one")
                    row += 1
                row += 1
            row += 1

        for r in range(1, row):
            ws.row_dimensions[r].height = 18

    def build_dashboard_prod_fam_wape(ws_name: str, df: pd.DataFrame) -> None:
        if ws_name in wb.sheetnames:
            del wb[ws_name]
        ws = wb.create_sheet(ws_name, 2)
        ws.title = ws_name
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18

        row = 1
        ws.cell(row=row, column=1, value="Forecast Accuracy Dashboard - Product Family (WAPE)").font = title_font
        row += 2

        bu_order = {"D100": 0, "D200": 1, "D300": 2}
        for (bu_code, bu_name) in (
            df[["BU Code", "BU Name"]]
            .drop_duplicates()
            .assign(_order=lambda d: d["BU Code"].map(bu_order).fillna(99))
            .sort_values(["_order", "BU Code"])
            .drop(columns=["_order"])
            .itertuples(index=False, name=None)
        ):
            bu_df = df[(df["BU Code"] == bu_code) & (df["BU Name"] == bu_name)]
            if bu_df.empty:
                continue
            write_section_header(ws, row, f"{bu_name} ({bu_code})", col_span=4)
            row += 2
            for units in ["quantity", "dollars"]:
                unit_df = bu_df[bu_df["Units"] == units]
                if unit_df.empty:
                    continue
                ws.cell(row=row, column=1, value=units.title()).font = header_font
                row += 1
                write_table_header(ws, row, ["Prod Fam", "Metric", "Stats Model", "Marketing"])
                row += 1
                prod_fams = unit_df["Prod Fam"].dropna().unique().tolist()
                sort_rows = []
                for prod_fam in prod_fams:
                    pf_df = unit_df[(unit_df["Prod Fam"] == prod_fam) & (unit_df["Metric"] == "WAPE")]
                    if pf_df.empty:
                        continue
                    stats_val = pf_df["Stats Model"].iloc[0]
                    mkt_val = pf_df["Marketing"].iloc[0]
                    winner = winner_label(stats_val, mkt_val, mode="lower")
                    winner_rank = 2
                    if winner == "Stats Model":
                        winner_rank = 0
                    elif winner == "Marketing":
                        winner_rank = 1
                    sort_rows.append((winner_rank, str(prod_fam), stats_val, mkt_val))

                for _, prod_fam, stats_val, mkt_val in sorted(sort_rows, key=lambda r: (r[0], r[1])):
                    write_metric_row(
                        ws,
                        row,
                        [prod_fam, "WAPE", stats_val, mkt_val],
                        percent_cols={3, 4},
                    )
                    highlight_winner(ws, row, 3, 4, mode="lower")
                    row += 1
                row += 1
            row += 1

        for r in range(1, row):
            ws.row_dimensions[r].height = 18

    def build_dashboard_marketing_mgr(ws_name: str, df: pd.DataFrame) -> None:
        if ws_name in wb.sheetnames:
            del wb[ws_name]
        ws = wb.create_sheet(ws_name, 3)
        ws.title = ws_name
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

        row = 1
        ws.cell(row=row, column=1, value="Forecast Accuracy Dashboard - Marketing Manager").font = title_font
        row += 2

        bu_order = {"D100": 0, "D200": 1, "D300": 2}
        for (bu_code, bu_name) in (
            df[["BU Code", "BU Name"]]
            .drop_duplicates()
            .assign(_order=lambda d: d["BU Code"].map(bu_order).fillna(99))
            .sort_values(["_order", "BU Code"])
            .drop(columns=["_order"])
            .itertuples(index=False, name=None)
        ):
            bu_df = df[(df["BU Code"] == bu_code) & (df["BU Name"] == bu_name)]
            if bu_df.empty:
                continue
            write_section_header(ws, row, f"{bu_name} ({bu_code})", col_span=4)
            row += 2
            for units in ["quantity", "dollars"]:
                unit_df = bu_df[bu_df["Units"] == units]
                if unit_df.empty:
                    continue
                ws.cell(row=row, column=1, value=units.title()).font = header_font
                row += 1
                write_table_header(ws, row, ["Marketing Manager", "Metric", "Stats Model", "Marketing"])
                row += 1

                managers = unit_df["Marketing Manager"].dropna().unique().tolist()
                sort_rows = []
                for manager in managers:
                    mgr_df = unit_df[(unit_df["Marketing Manager"] == manager) & (unit_df["Metric"] == "Midmark Fcast Acc.")]
                    if mgr_df.empty:
                        continue
                    stats_val = mgr_df["Stats Model"].iloc[0]
                    mkt_val = mgr_df["Marketing"].iloc[0]
                    winner = winner_label(stats_val, mkt_val, mode="closest_to_one")
                    winner_rank = 2
                    if winner == "Stats Model":
                        winner_rank = 0
                    elif winner == "Marketing":
                        winner_rank = 1
                    sort_rows.append((winner_rank, str(manager), stats_val, mkt_val))

                for _, manager, stats_val, mkt_val in sorted(sort_rows, key=lambda r: (r[0], r[1])):
                    write_metric_row(
                        ws,
                        row,
                        [manager, "Midmark Fcast Acc.", stats_val, mkt_val],
                        percent_cols={3, 4},
                    )
                    highlight_winner(ws, row, 3, 4, mode="closest_to_one")
                    row += 1
                row += 1
            row += 1

        for r in range(1, row):
            ws.row_dimensions[r].height = 18

    def build_dashboard_product(ws_name: str, df: pd.DataFrame) -> None:
        if ws_name in wb.sheetnames:
            del wb[ws_name]
        ws = wb.create_sheet(ws_name, 4)
        ws.title = ws_name
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

        row = 1
        ws.cell(row=row, column=1, value="Forecast Accuracy Dashboard - Product").font = title_font
        row += 2

        bu_order = {"D100": 0, "D200": 1, "D300": 2}
        for (bu_code, bu_name) in (
            df[["BU Code", "BU Name"]]
            .drop_duplicates()
            .assign(_order=lambda d: d["BU Code"].map(bu_order).fillna(99))
            .sort_values(["_order", "BU Code"])
            .drop(columns=["_order"])
            .itertuples(index=False, name=None)
        ):
            bu_df = df[(df["BU Code"] == bu_code) & (df["BU Name"] == bu_name)]
            if bu_df.empty:
                continue
            write_section_header(ws, row, f"{bu_name} ({bu_code})", col_span=4)
            row += 2
            for units in ["quantity", "dollars"]:
                unit_df = bu_df[bu_df["Units"] == units]
                if unit_df.empty:
                    continue
                ws.cell(row=row, column=1, value=units.title()).font = header_font
                row += 1
                write_table_header(ws, row, ["Product", "Marketing Manager", "Stats Model", "Marketing"])
                row += 1

                products = unit_df["Product"].dropna().unique().tolist()
                sort_rows = []
                for product in products:
                    prod_df = unit_df[
                        (unit_df["Product"] == product) & (unit_df["Metric"] == "Midmark Fcast Acc.")
                    ]
                    if prod_df.empty:
                        continue
                    stats_val = prod_df["Stats Model"].iloc[0]
                    mkt_val = prod_df["Marketing"].iloc[0]
                    manager = prod_df["Marketing Manager"].iloc[0]

                    product_blank = pd.isna(product) or str(product).strip() == ""
                    manager_blank = pd.isna(manager) or str(manager).strip() == ""
                    stats_blank = pd.isna(stats_val)
                    mkt_blank = pd.isna(mkt_val)
                    hidden = (
                        product_blank
                        or manager_blank
                        or stats_blank
                        or mkt_blank
                        or (stats_val == 0 and mkt_val == 0)
                    )

                    winner = winner_label(stats_val, mkt_val, mode="closest_to_one")
                    winner_rank = 2
                    if winner == "Stats Model":
                        winner_rank = 0
                    elif winner == "Marketing":
                        winner_rank = 1

                    manager_sort = str(manager).strip() if not manager_blank else "ZZZZZZ"
                    product_sort = str(product).strip() if not product_blank else "ZZZZZZ"
                    sort_rows.append(
                        (winner_rank, manager_sort, product_sort, str(product), str(manager), stats_val, mkt_val, hidden)
                    )

                for _, _, _, product, manager, stats_val, mkt_val, hidden in sorted(
                    sort_rows, key=lambda r: (r[0], r[1], r[2])
                ):
                    write_metric_row(
                        ws,
                        row,
                        [product, manager, stats_val, mkt_val],
                        percent_cols={3, 4},
                    )
                    if not hidden:
                        highlight_winner(ws, row, 3, 4, mode="closest_to_one")
                    else:
                        ws.row_dimensions[row].hidden = True
                    row += 1
                row += 1
            row += 1

        for r in range(1, row):
            ws.row_dimensions[r].height = 18

    build_dashboard_totals("Dashboard - Totals", totals)
    build_dashboard_prod_fam("Dashboard - Prod Fam", prod_fam)
    build_dashboard_prod_fam_wape("Dashboard - Prod Fam WAPE", prod_fam_wape)
    build_dashboard_marketing_mgr("Dashboard - Mkt Mgr", marketing_mgr)
    build_dashboard_product("Dashboard - Product", product)

    for sheet_name in [
        "Dashboard - Mkt Mgr",
        "Dashboard - Prod Fam WAPE",
        "Dashboard - Totals Data",
        "Dashboard - Prod Fam Data",
        "Dashboard - Prod Fam WAPE Data",
        "Dashboard - Mkt Mgr Data",
        "Dashboard - Product Data",
        "Raw Data",
        "Skipped Products",
        "Lookup",
    ]:
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_state = "hidden"

    wb.save(config.output_file)

    # Validation workbook
    validation_path = config.output_file.parent / f"{month_label(config.report_month)} Forecast Accuracy Validation.xlsx"
    with pd.ExcelWriter(validation_path, engine="openpyxl") as writer:
        marketing_detail.to_excel(writer, sheet_name="Marketing Mapped", index=False)
        marketing_rollup.to_excel(writer, sheet_name="Marketing Rollup", index=False)
        actuals_rollup.to_excel(writer, sheet_name="Actuals Rollup", index=False)
        stats_rollup.to_excel(writer, sheet_name="Stats Rollup (BLEND)", index=False)
        raw.to_excel(writer, sheet_name="Raw Data", index=False)
        totals.to_excel(writer, sheet_name="Totals Data", index=False)
        prod_fam.to_excel(writer, sheet_name="Prod Fam Data", index=False)
        prod_fam_wape.to_excel(writer, sheet_name="Prod Fam WAPE Data", index=False)
        product.to_excel(writer, sheet_name="Product Data", index=False)


def main() -> None:
    args = parse_args()
    report_month = resolve_report_month(args.month)
    stats_file = ROOT / stats_model_filename(report_month)
    if not stats_file.exists():
        raise SystemExit(f"Stats model file not found: {stats_file}")

    output_name = args.output or f"{month_label(report_month)} Forecast Accuracy Report.xlsx"
    output_path = ROOT / output_name

    config = ReportConfig(
        report_month=report_month,
        marketing_file=MARKETING_FILE,
        product_catalog_file=PRODUCT_CATALOG_FILE,
        stats_model_file=stats_file,
        output_file=output_path,
    )

    write_report(config)
    print(f"Report written: {output_path}")


if __name__ == "__main__":
    main()
