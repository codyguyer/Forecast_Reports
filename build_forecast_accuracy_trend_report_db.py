from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill

import build_forecast_accuracy_report as legacy
import build_forecast_accuracy_report_db as db_report


ROOT = Path(__file__).resolve().parent


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build rolling forecast-accuracy trend report from DB sources."
    )
    parser.add_argument(
        "--month",
        type=str,
        default=None,
        help="Anchor month in YYYY-MM. Defaults to previous month (last fully closed month).",
    )
    parser.add_argument("--window-months", type=int, default=12)
    parser.add_argument("--top-n-products", type=int, default=10)
    parser.add_argument("--server", type=str, required=True)
    parser.add_argument("--database", type=str, default="Forecast_Database")
    parser.add_argument("--driver", type=str, default="ODBC Driver 17 for SQL Server")
    parser.add_argument("--dq-mode", choices=["off", "warn", "fail"], default="fail")
    parser.add_argument("--dq-log", type=str, default=None)
    parser.add_argument("--output", type=str, default=None)
    return parser.parse_args()


def month_sequence(start_month: date, end_month: date) -> list[date]:
    months: list[date] = []
    current = pd.Timestamp(start_month)
    end = pd.Timestamp(end_month)
    while current <= end:
        months.append(current.date())
        current = current + pd.offsets.MonthBegin(1)
    return months


def resolve_window(anchor_month: date, window_months: int) -> tuple[date, date]:
    if window_months <= 0:
        raise SystemExit("--window-months must be >= 1")
    anchor_ts = pd.Timestamp(anchor_month)
    start_ts = anchor_ts - pd.offsets.MonthBegin(window_months - 1)
    return start_ts.date(), anchor_month


def safe_ratio(numerator: float, denominator: float) -> float | None:
    if denominator == 0:
        return None
    return float(numerator) / float(denominator)


def _view_rows(
    frame: pd.DataFrame,
    month_start: date,
    units: str,
    view_level: str,
    bu_code: str,
    bu_name: str,
    prod_fam: str,
    product: str,
    marketing_manager: str,
) -> list[dict[str, Any]]:
    actuals_sum = float(frame["Actuals"].sum())
    stats_forecast_sum = float(frame["Stats Model Fcast"].sum())
    marketing_forecast_sum = float(frame["Marketing Fcast"].sum())
    stats_abs_error_sum = float(frame["Stats Abs Error"].sum())
    marketing_abs_error_sum = float(frame["Marketing Abs Error"].sum())

    stats_acc = safe_ratio(actuals_sum, stats_forecast_sum)
    marketing_acc = safe_ratio(actuals_sum, marketing_forecast_sum)

    base = {
        "month_start": pd.Timestamp(month_start),
        "units": units,
        "view_level": view_level,
        "bu_code": bu_code,
        "bu_name": bu_name,
        "prod_fam": prod_fam,
        "product": product,
        "marketing_manager": marketing_manager,
        "metric_name": "midmark_fcast_acc",
        "actuals_sum": actuals_sum,
    }

    return [
        {
            **base,
            "model_side": "stats_model",
            "metric_value": stats_acc,
            "forecast_sum": stats_forecast_sum,
            "abs_error_sum": stats_abs_error_sum,
        },
        {
            **base,
            "model_side": "marketing",
            "metric_value": marketing_acc,
            "forecast_sum": marketing_forecast_sum,
            "abs_error_sum": marketing_abs_error_sum,
        },
    ]


def build_month_trend_rows(raw: pd.DataFrame, month_start: date) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    for units in ["quantity", "dollars"]:
        unit_df = raw[raw["Units"] == units]
        if unit_df.empty:
            continue

        records.extend(
            _view_rows(
                frame=unit_df,
                month_start=month_start,
                units=units,
                view_level="total",
                bu_code="ALL",
                bu_name="Total",
                prod_fam="",
                product="",
                marketing_manager="",
            )
        )

        for (bu_code, bu_name), group in unit_df.groupby(["Division", "BU Name"], dropna=False):
            records.extend(
                _view_rows(
                    frame=group,
                    month_start=month_start,
                    units=units,
                    view_level="bu",
                    bu_code=str(bu_code),
                    bu_name=str(bu_name),
                    prod_fam="",
                    product="",
                    marketing_manager="",
                )
            )

        for (bu_code, bu_name, prod_fam), group in unit_df.groupby(
            ["Division", "BU Name", "Prod Fam"], dropna=False
        ):
            records.extend(
                _view_rows(
                    frame=group,
                    month_start=month_start,
                    units=units,
                    view_level="prod_fam",
                    bu_code=str(bu_code),
                    bu_name=str(bu_name),
                    prod_fam=str(prod_fam),
                    product="",
                    marketing_manager="",
                )
            )

        for (bu_code, bu_name, product, manager), group in unit_df.groupby(
            ["Division", "BU Name", "Product", "Marketing Manager"], dropna=False
        ):
            records.extend(
                _view_rows(
                    frame=group,
                    month_start=month_start,
                    units=units,
                    view_level="product",
                    bu_code=str(bu_code),
                    bu_name=str(bu_name),
                    prod_fam=str(group["Prod Fam"].iloc[0]) if not group.empty else "",
                    product=str(product),
                    marketing_manager=str(manager),
                )
            )

    return pd.DataFrame.from_records(records)


def run_dq_checks(
    trend_df: pd.DataFrame,
    expected_months: list[date],
    source_row_counts: list[dict[str, Any]],
    top_n_df: pd.DataFrame,
    top_n_products: int,
) -> dict[str, Any]:
    checks: list[dict[str, Any]] = []

    def add_check(name: str, severity: str, passed: bool, details: dict[str, Any]) -> None:
        checks.append(
            {
                "name": name,
                "severity": severity,
                "passed": bool(passed),
                "details": details,
            }
        )

    available_months = sorted({d.date() for d in pd.to_datetime(trend_df["month_start"], errors="coerce").dropna()})
    expected_set = set(expected_months)
    available_set = set(available_months)
    missing_months = sorted(expected_set - available_set)
    add_check(
        name="rolling_window_completeness",
        severity="warning",
        passed=len(missing_months) == 0,
        details={
            "expected_months": [str(m) for m in expected_months],
            "available_months": [str(m) for m in available_months],
            "missing_months": [str(m) for m in missing_months],
            "expected_count": len(expected_months),
            "available_count": len(available_months),
        },
    )

    coverage_issues = []
    for c in source_row_counts:
        if c["marketing_rows"] == 0 or c["stats_rows"] == 0 or c["actuals_rows"] == 0:
            coverage_issues.append(c)
    add_check(
        name="source_coverage_by_month",
        severity="critical",
        passed=len(coverage_issues) == 0,
        details={"issues": coverage_issues[:20], "issue_count": len(coverage_issues)},
    )

    dupes = trend_df.duplicated(
        subset=[
            "month_start",
            "units",
            "view_level",
            "bu_code",
            "bu_name",
            "prod_fam",
            "product",
            "marketing_manager",
            "model_side",
        ]
    ).sum()
    add_check(
        name="trend_grain_uniqueness",
        severity="critical",
        passed=int(dupes) == 0,
        details={"duplicate_count": int(dupes)},
    )

    invalid_denominator = trend_df[(trend_df["forecast_sum"] <= 0) & (trend_df["metric_value"].notna())]
    add_check(
        name="denominator_validity",
        severity="critical",
        passed=invalid_denominator.empty,
        details={
            "invalid_count": int(len(invalid_denominator)),
            "examples": invalid_denominator.head(10).to_dict(orient="records"),
        },
    )

    topn_counts = (
        top_n_df[["units", "product"]]
        .drop_duplicates()
        .groupby("units", dropna=False)["product"]
        .nunique()
        .reset_index(name="topn_count")
    )
    over_limit = topn_counts[topn_counts["topn_count"] > top_n_products]
    add_check(
        name="topn_rank_integrity",
        severity="critical",
        passed=over_limit.empty,
        details={
            "top_n_products": int(top_n_products),
            "counts": topn_counts.to_dict(orient="records"),
            "over_limit": over_limit.to_dict(orient="records"),
        },
    )

    summary = {
        "checks_total": len(checks),
        "checks_failed": int(sum(0 if c["passed"] else 1 for c in checks)),
        "critical_failed": int(sum(0 if (c["passed"] or c["severity"] != "critical") else 1 for c in checks)),
        "warning_failed": int(sum(0 if (c["passed"] or c["severity"] != "warning") else 1 for c in checks)),
        "checks": checks,
    }
    return summary


def build_views(trend_df: pd.DataFrame, anchor_month: date, top_n_products: int) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    totals = trend_df[trend_df["view_level"] == "total"].copy()
    totals = totals[["month_start", "units", "model_side", "metric_value"]]
    totals_wide = totals.pivot_table(
        index=["month_start", "units"], columns="model_side", values="metric_value", aggfunc="first"
    ).reset_index()
    totals_wide["delta_stats_minus_marketing"] = totals_wide["stats_model"] - totals_wide["marketing"]
    totals_wide = totals_wide.rename(columns={"stats_model": "stats_model_acc", "marketing": "marketing_acc"})

    bu = trend_df[trend_df["view_level"] == "bu"].copy()
    bu_wide = bu.pivot_table(
        index=["month_start", "units", "bu_code", "bu_name"],
        columns="model_side",
        values="metric_value",
        aggfunc="first",
    ).reset_index()
    bu_wide["delta_stats_minus_marketing"] = bu_wide["stats_model"] - bu_wide["marketing"]
    bu_wide = bu_wide.rename(columns={"stats_model": "stats_model_acc", "marketing": "marketing_acc"})

    prod_fam = trend_df[trend_df["view_level"] == "prod_fam"].copy()
    prod_fam_wide = prod_fam.pivot_table(
        index=["month_start", "units", "bu_code", "bu_name", "prod_fam"],
        columns="model_side",
        values="metric_value",
        aggfunc="first",
    ).reset_index()
    prod_fam_wide["delta_stats_minus_marketing"] = prod_fam_wide["stats_model"] - prod_fam_wide["marketing"]
    prod_fam_wide = prod_fam_wide.rename(columns={"stats_model": "stats_model_acc", "marketing": "marketing_acc"})

    product_all = trend_df[trend_df["view_level"] == "product"].copy()
    latest = product_all[pd.to_datetime(product_all["month_start"]).dt.date == anchor_month].copy()
    latest_rank = (
        latest[["units", "product", "actuals_sum"]]
        .drop_duplicates()
        .sort_values(["units", "actuals_sum", "product"], ascending=[True, False, True])
    )
    latest_rank["rank"] = latest_rank.groupby("units").cumcount() + 1
    top_keys = latest_rank[latest_rank["rank"] <= top_n_products][["units", "product"]]

    product_top = product_all.merge(top_keys, on=["units", "product"], how="inner")
    product_top_wide = product_top.pivot_table(
        index=[
            "month_start",
            "units",
            "bu_code",
            "bu_name",
            "prod_fam",
            "product",
            "marketing_manager",
            "actuals_sum",
        ],
        columns="model_side",
        values="metric_value",
        aggfunc="first",
    ).reset_index()
    product_top_wide["delta_stats_minus_marketing"] = product_top_wide["stats_model"] - product_top_wide["marketing"]
    product_top_wide = product_top_wide.rename(columns={"stats_model": "stats_model_acc", "marketing": "marketing_acc"})

    return totals_wide, bu_wide, prod_fam_wide, product_top_wide


def apply_workbook_formatting(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, size=12)
    header_font = Font(bold=True)

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        if ws.max_row == 0 or ws.max_column == 0:
            continue
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
        ws.freeze_panes = "A2"

    exec_ws = wb["Trend - Executive"]
    exec_ws.insert_rows(1)
    exec_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(8, exec_ws.max_column))
    exec_ws.cell(row=1, column=1, value="Forecast Accuracy Trend - Executive View")
    exec_ws.cell(row=1, column=1).fill = title_fill
    exec_ws.cell(row=1, column=1).font = title_font

    exec_ws.freeze_panes = "A3"

    def add_chart(units_value: str, start_row: int, title: str) -> None:
        data_rows = [
            r for r in range(3, exec_ws.max_row + 1) if str(exec_ws.cell(row=r, column=2).value).lower() == units_value
        ]
        if not data_rows:
            return
        min_r, max_r = min(data_rows), max(data_rows)

        chart = LineChart()
        chart.title = title
        chart.y_axis.title = "Accuracy"
        chart.x_axis.title = "Month"
        chart.height = 6
        chart.width = 10

        # stats_model_acc column C, marketing_acc column D
        data = Reference(exec_ws, min_col=3, max_col=4, min_row=min_r, max_row=max_r)
        cats = Reference(exec_ws, min_col=1, min_row=min_r, max_row=max_r)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        chart.style = 2
        exec_ws.add_chart(chart, f"H{start_row}")

    add_chart("quantity", 3, "Quantity Trend: Stats vs Marketing")
    add_chart("dollars", 20, "Dollars Trend: Stats vs Marketing")

    wb.save(workbook_path)


def write_outputs(
    output_path: Path,
    trend_df: pd.DataFrame,
    totals_df: pd.DataFrame,
    bu_df: pd.DataFrame,
    prod_fam_df: pd.DataFrame,
    product_top_df: pd.DataFrame,
    dq_summary: dict[str, Any],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    qa_rows: list[dict[str, Any]] = []
    for c in dq_summary["checks"]:
        qa_rows.append(
            {
                "check_name": c["name"],
                "severity": c["severity"],
                "passed": c["passed"],
                "details": json.dumps(c["details"], default=str),
            }
        )
    qa_df = pd.DataFrame.from_records(qa_rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        totals_df.to_excel(writer, sheet_name="Trend - Executive", index=False)
        bu_df.to_excel(writer, sheet_name="Trend - BU", index=False)
        prod_fam_df.to_excel(writer, sheet_name="Trend - Prod Fam", index=False)
        product_top_df.to_excel(writer, sheet_name="Trend - Product Top10", index=False)
        trend_df.sort_values(
            ["month_start", "units", "view_level", "bu_code", "prod_fam", "product", "model_side"]
        ).to_excel(writer, sheet_name="Trend - Data", index=False)
        qa_df.to_excel(writer, sheet_name="Trend - QA", index=False)

    apply_workbook_formatting(output_path)


def main() -> None:
    args = parse_args()
    anchor_month = legacy.resolve_report_month(args.month)
    start_month, end_month = resolve_window(anchor_month, args.window_months)
    months = month_sequence(start_month, end_month)

    sql_cfg = db_report.SqlConfig(server=args.server, database=args.database, driver=args.driver)
    output_name = args.output or f"outputs/reports/{legacy.month_label(anchor_month)} Forecast Accuracy Trend Report.xlsx"
    output_path = ROOT / output_name
    dq_log = Path(args.dq_log) if args.dq_log else ROOT / "outputs" / "comparisons" / (
        f"{legacy.month_label(anchor_month)} Forecast Accuracy Trend DQ (db).json"
    )

    source_counts: list[dict[str, Any]] = []
    trend_frames: list[pd.DataFrame] = []

    for month in months:
        marketing_df, catalog_df, stats_df, actuals_df = db_report.load_from_db(month, sql_cfg)
        source_counts.append(
            {
                "month_start": str(month),
                "marketing_rows": int(len(marketing_df)),
                "catalog_rows": int(len(catalog_df)),
                "stats_rows": int(len(stats_df)),
                "actuals_rows": int(len(actuals_df)),
            }
        )

        if marketing_df.empty or stats_df.empty or actuals_df.empty:
            continue

        raw, _, _, _, _, _ = legacy.build_raw_data(
            marketing_df=marketing_df,
            catalog_df=catalog_df,
            stats_df=stats_df,
            actuals_df=actuals_df,
            report_month=month,
        )
        if raw.empty:
            continue
        trend_frames.append(build_month_trend_rows(raw, month))

    if not trend_frames:
        raise SystemExit("No trend data rows were generated for the selected rolling window.")

    trend_df = pd.concat(trend_frames, ignore_index=True)
    totals_df, bu_df, prod_fam_df, product_top_df = build_views(
        trend_df=trend_df,
        anchor_month=anchor_month,
        top_n_products=args.top_n_products,
    )

    dq_summary = run_dq_checks(
        trend_df=trend_df,
        expected_months=months,
        source_row_counts=source_counts,
        top_n_df=product_top_df,
        top_n_products=args.top_n_products,
    )
    dq_summary.update(
        {
            "anchor_month": str(anchor_month),
            "window_start": str(start_month),
            "window_end": str(end_month),
            "window_months": int(args.window_months),
            "top_n_products": int(args.top_n_products),
            "source_counts": source_counts,
            "output": str(output_path),
        }
    )

    dq_log.parent.mkdir(parents=True, exist_ok=True)
    dq_log.write_text(json.dumps(dq_summary, indent=2, default=str), encoding="utf-8")

    write_outputs(
        output_path=output_path,
        trend_df=trend_df,
        totals_df=totals_df,
        bu_df=bu_df,
        prod_fam_df=prod_fam_df,
        product_top_df=product_top_df,
        dq_summary=dq_summary,
    )

    if args.dq_mode == "warn" and dq_summary["checks_failed"] > 0:
        print(f"[DQ-WARN] Failed checks: {dq_summary['checks_failed']}; critical: {dq_summary['critical_failed']}")
    if args.dq_mode == "fail" and dq_summary["critical_failed"] > 0:
        raise SystemExit(
            f"[DQ-FAIL] Critical checks failed: {dq_summary['critical_failed']}. See {dq_log}"
        )

    print(f"[DONE] Trend workbook written: {output_path}")
    print(f"[DONE] Trend DQ log written: {dq_log}")


if __name__ == "__main__":
    main()
